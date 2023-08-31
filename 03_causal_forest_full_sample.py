# %%
import pandas as pd
import numpy as np

from sklearn.model_selection import train_test_split
from econml.grf import CausalForest
from library import start
import statsmodels.formula.api as smf


import matplotlib.pyplot as plt
from sklearn.tree import plot_tree
from openpyxl import load_workbook

# %%
FILE_NAME = "feedback_analysis_withpre_post_survey_wide.dta"

PREDICTORS = [
    "das_stress",
    "das_depression",
    "das_anxiety",
    "dass_total",
    "neo_n",
    "neo_e",
    "neo_a",
    "neo_c",
    "neo_o",
    "tses_is",
    "score0",
    "score1",
    "elem",
    # "rsq_total",
]
df = pd.read_stata(
    start.RAW_DATA_DIR + FILE_NAME,
)
df["treat"] = np.where(df.treat == "Coaching", 1, 0)
df.treat.value_counts()
# %%

treatment = "treat"
outcome = "score2"
covariates = PREDICTORS
all_variables = [treatment] + [outcome] + covariates
df = df.dropna(axis=0, subset=all_variables)
len(df)
# %%

Y = df[outcome]
T = df[treatment]
X = df[covariates]


# %%

causal_forest = CausalForest(
    n_estimators=1000,
    criterion="het",
    max_depth=None,
    min_samples_split=10,
    min_samples_leaf=5,
    max_samples=0.95,
    honest=True,
    inference=False,
    random_state=11,
)

causal_forest.fit(X=X, T=T, y=Y)

# %%
forest_importances = pd.Series(causal_forest.feature_importances_, index=PREDICTORS)
forest_importances = pd.DataFrame(forest_importances).sort_values(by=0, ascending=False)
forest_importances
# %%
df["itt_prediction"] = causal_forest.predict(X=X)
df.itt_prediction.hist()
# %%

# %%

df["itt_quartile"] = pd.qcut(df["itt_prediction"], q=4, labels=False)


# %%

RESULTS_FILE = start.MAIN_DIR + "results/" + "HTE Characteristics.xlsx"
file = RESULTS_FILE
wb = load_workbook(file)
ws = wb["raw"]

characteristics = ["neo_o", "dass_total", "tses_is"]

row = 2
for char in characteristics:
    col = 2
    mean_value = df[df.itt_quartile == 3][char].mean().round(2)
    ws.cell(row=row, column=col).value = mean_value

    col = 3
    mean_value = df[df.itt_quartile == 0][char].mean().round(2)
    ws.cell(row=row, column=col).value = mean_value

    col = 4
    mean_value = df[char].mean().round(2)
    ws.cell(row=row, column=col).value = mean_value

    row = row + 1


for char in characteristics:
    col = 2
    mean_value = df[df.itt_quartile == 3].itt_prediction.mean().round(2)
    ws.cell(row=6, column=col).value = mean_value
    mod = smf.ols(
        formula="score2 ~ treat + C(strata)",
        data=df[df.itt_quartile == 3],
    )
    res = mod.fit()
    coef = res.params["treat"].round(2)
    ws.cell(row=7, column=col).value = coef

    col = 3
    mean_value = df[df.itt_quartile == 0].itt_prediction.mean().round(2)
    ws.cell(row=6, column=col).value = mean_value
    mod = smf.ols(
        formula="score2 ~ treat + C(strata)",
        data=df[df.itt_quartile == 0],
    )
    res = mod.fit()
    coef = res.params["treat"].round(2)
    ws.cell(row=7, column=col).value = coef

    col = 4
    mean_value = df.itt_prediction.mean().round(2)
    ws.cell(row=6, column=col).value = mean_value
    mod = smf.ols(
        formula="score2 ~ treat + C(strata)",
        data=df,
    )
    res = mod.fit()
    coef = res.params["treat"].round(2)
    ws.cell(row=7, column=col).value = coef

    row = row + 1

wb.save(file)
# %%


causal_tree = CausalForest(
    n_estimators=1,
    criterion="het",
    max_depth=2,
    min_samples_split=10,
    min_samples_leaf=5,
    max_samples=0.99,
    honest=True,
    inference=False,
    random_state=11,
)

causal_tree.fit(X=df[characteristics], T=T, y=Y)
plt.figure(figsize=(20, 10))
plot_tree(causal_tree[0], impurity=True, max_depth=3)
plt.show()
characteristics
# %%
